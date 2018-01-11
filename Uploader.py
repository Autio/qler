import pytumblr
import os
import xlsxwriter
import openpyxl
from random import shuffle

# Read master Excel file into list
from bs4 import BeautifulSoup
import requests
import re
import urllib2
import os
import xlsxwriter
import cookielib
import json
import csv
import sys
reload(sys)
sys.setdefaultencoding('utf8')
printDesc = False

def get_soup(url,header):
    return BeautifulSoup(urllib2.urlopen(urllib2.Request(url,headers=header)),'html.parser')

def get_files(searchTerm = [], destinationFolder = ""):
    imageList = []
    query = searchTerm # you can change the query for the image  here
    image_type="ActiOn"
    query= query.split()
    query='+'.join(query)
    url="https://www.google.co.in/search?q="+query+"&source=lnms&tbm=isch"
    print url
    #add the directory for your image here
    DIR=destinationFolder
    header={'User-Agent':"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/43.0.2357.134 Safari/537.36"
    }
    soup = get_soup(url,header)

    ActualImages=[]# contains the link for Large original images, type of  image
    for a in soup.find_all("div",{"class":"rg_meta"}):
        link , Type =json.loads(a.text)["ou"]  ,json.loads(a.text)["ity"]
        desc = json.loads(a.text)["s"]
        desc = desc.replace('\\u', '')
        ActualImages.append([link,Type,desc])

    print  "there are total" , len(ActualImages),"images"

    if not os.path.exists(DIR):
                os.mkdir(DIR)
    DIR = os.path.join(DIR, query.split()[0])

    if not os.path.exists(DIR):
                os.mkdir(DIR)
    ###print images
    for i , [img, Type, desc] in enumerate(ActualImages):
        try:
            req = urllib2.Request(img, headers={'User-Agent' : header})
            raw_img = urllib2.urlopen(req).read()

            cntr = len([i for i in os.listdir(DIR) if image_type in i]) + 1
            print cntr
            if len(Type)==0:
                f = open(os.path.join(DIR , image_type + "_"+ str(cntr)+".jpg"), 'wb')
                imageList.append([image_type + "_" + str(cntr)+".jpg", img.encode('utf-8'), desc.encode('utf-8')])

            else :
                f = open(os.path.join(DIR , image_type + "_"+ str(cntr)+"."+Type), 'wb')
                imageList.append([image_type + "_"+ str(cntr)+"."+Type, img.encode('utf-8'), desc.encode('utf-8')])


            f.write(raw_img)
            f.close()
        except Exception as e:
            print "could not load : "+img
            print e

    print imageList
    workbook = xlsxwriter.Workbook(DIR + '\\index.xlsx')
    worksheet = workbook.add_worksheet()
    for row in range(len(imageList)):
        worksheet.write(row, 0, imageList[row][0])
        worksheet.write(row , 1, imageList[row][1])
        worksheet.write(row , 2, imageList[row][2])
    workbook.close()

#with open(DIR + '\\index.csv', 'wb') as myfile:
#    wr = csv.writer(myfile, delimiter='|', quoting=csv.QUOTE_ALL)
#    wr.writerow(row)

class TumblrBlog(object):
    def __init__(self, id, name, password, consumerKey, consumerSecret, oauthKey, oauthSecret, tags, keywords, printDescFlag):
        self.id = id
        self.name = name
        self.password = password
        self.consumerKey = consumerKey
        self.consumerSecret = consumerSecret
        self.oauthKey = oauthKey
        self.oauthSecret = oauthSecret
        self.tags = tags
        self.keywords = keywords
        self.printDescFlag = printDescFlag


tumblrData = []
# Read index file
mainFile = openpyxl.load_workbook("control file.xlsx")
mainSheet = mainFile['Sheet1']
headerRow = True
for row in mainSheet.rows:
    if not headerRow:
        line = []
        for cell in row:
            line.append(cell.value)
        if line[10] != "":
            tumblrData.append(TumblrBlog(line[0],line[1],line[4],line[5],line[6],line[7],line[8], line[9].split(','), line[10], line[11]))
    else:
        headerRow = False

count = 1
# cycle through each row
for tumblr in tumblrData:
    saveDir = os.getcwd() + "\\Temp\\"

    # Wipe existing?

    # Download based keywords
    get_files(tumblr.keywords, saveDir)

    # read in keys
    consumer_key = tumblr.consumerKey
    consumer_secret = tumblr.consumerSecret
    oauth_key = tumblr.oauthKey
    oauth_secret = tumblr.oauthSecret

    client = pytumblr.TumblrRestClient(
        consumer_key,
        consumer_secret,
        oauth_key,
        oauth_secret
    )

    print(client.info())

    # Parameters for mass posting
    folder = saveDir
    blogname = tumblr.name
    tags = tumblr.tags
    query = tumblr.keywords.split()
    query = '+'.join(query)
    folder = folder + query + "\\"
    posting_files = os.listdir(folder)

    index = []

    # Read index file
    wb = openpyxl.load_workbook(folder + "index.xlsx")
    ws = wb['Sheet1']

    for row in ws.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        index.append(line)
    #with open(folder + 'index.csv', 'r') as myfile:
    #    for line in myfile:
    #        index.append(line.strip().split('|'))
    # Only output files which have a corresponding line in the index file

     #cycle through them all and post as queue
    counter = 1
    for f in posting_files:
        shuffle(tags)
        for line in index:
            try:
                if(line[0].strip('"') == f):
                    source = line[1]
                    description = ""
                    try:
                        description = line[2].strip("'").strip(',').strip('"')
                    except:
                        print 'no description for file ' + str(counter)
                   # Createds a photo post using a local filepath
                    print 'Uploading file ' + str(counter)
                    folder = str(folder)
                    f = str(f)
                    printDesc = tumblr.printDescFlag
                    if not printDesc:
                        description = ""
                    client.create_photo(blogname, state="queue", tags=tags, caption= description + "\n\n<a href = \"" + source.replace('"', '') + "\">Source</a>\n\n", data=folder + f)
                    print 'File ' + f + ' uploaded'
            except Exception, e:
                print 'Failed to upload to ftp: ' + str(e)
                print 'Could not upload file ' + str(counter)
        counter += 1
